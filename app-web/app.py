from flask import Flask, render_template, request, redirect, url_for, session, make_response, send_from_directory, send_file, jsonify
from flask_mysqldb import MySQL, MySQLdb
from werkzeug.utils import secure_filename
import pandas as pd

from datetime import datetime, timedelta
from weasyprint import HTML
import os
from babel.dates import format_date
import bcrypt
from markupsafe import Markup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import subprocess
from functools import wraps

app = Flask(__name__)
# Configuraciones
app.config['SECRET_KEY'] = '3054=HitM'
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'user623'
app.config['MYSQL_DB'] = 'abrilpasivoss'
MySQL = MySQL(app)
# app.config['SESSION_TYPE'] = 'filesystem' 
# app.config['SESSION_PERMANENT'] = False
# app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=15)  
# Session(app)
# from flask_session import Session

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(os.path.join(app.root_path, 'static'), filename)

@app.route("/logout")
def logout():
    cedula = session.get('cedula')
    username = session.get('username')
    time_login = session.get('time_login')
    time_finish = datetime.now()
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('INSERT INTO user_history (cedula, user, action, time_login,time_finish) VALUES (%s, %s, %s, %s,%s)',(cedula, username,'sesion cerrada',time_login,time_finish))
    MySQL.connection.commit()
    cursor.close()
    session.pop('loggedin', None)
    session.pop('cedula', None)
    session.pop('username', None)
    session.pop('Super_Admin', None) 
    return redirect(url_for('login'))



@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        cedula = request.form['cedula']
        password = request.form['password']
        
        if not cedula or not password:
            error = "Cédula y contraseña son requeridas"
            return render_template('login.html', error=error)

        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT * FROM tabla WHERE Cedula = %s', (cedula,))
        user = cursor.fetchone()

        if user:
          
            if bcrypt.checkpw(password.encode('utf-8'), user['Password'].encode('utf-8')):
                if user['estado'] == 'suspendido':
                    error = "Tu cuenta está suspendida. Contacta al administrador."
                    return render_template('login.html', error=error)

                session['loggedin'] = True
                session['cedula'] = user['Cedula']
                session['username'] = user['username']
                session['Super_Admin'] = user['Super_Admin']
                session['time_login'] = datetime.now()
                cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', (session['cedula'], session['username'], 'sesion iniciada', datetime.now()))
                MySQL.connection.commit()
                cursor.close()
                if user['Super_Admin'] == 1:
                    return redirect(url_for('consult'))
                else:
                    return redirect(url_for('consult'))
            else:
                error = "Cédula o contraseña incorrecta"
                return render_template('login.html', error=error)
        else:
            error = "Cédula o contraseña incorrecta"
            return render_template('login.html', error=error)

    return render_template('login.html')


@app.route("/superAdmin")
def superAdmin():
    return render_template('superAdmin.html')

@app.route("/tipo_user", methods=["GET", "POST"])
def tipo_user():
    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == "POST":
        cedula = request.form['cedula']
        nuevo_estatus = request.form['super_admin']

        try:
            cursor.execute('SELECT COUNT(*) AS total_personas FROM tabla')
            total_personas = cursor.fetchone()['total_personas']
            cursor.execute(
                'UPDATE tabla SET Super_Admin = %s WHERE Cedula = %s',
                (nuevo_estatus, cedula)
            )
            MySQL.connection.commit()
        except Exception as e:
            MySQL.connection.rollback()
            return render_template("cambiar_super_admin.html", total_personas=total_personas, error=f"Error al actualizar: {str(e)}")

    cursor.execute('SELECT Cedula, username, Super_Admin FROM tabla')
    usuarios = cursor.fetchall()
    cursor.execute('SELECT COUNT(*) AS total_personas FROM tabla')
    total_personas = cursor.fetchone()['total_personas']
    cursor.close()

    return render_template("cambiar_super_admin.html", usuarios=usuarios, total_personas=total_personas)

@app.route("/RegistUser", methods=["GET", "POST"])
def RegistUser():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    if request.method == "POST":
        cedula = request.form['cedula']
        username = request.form['username']
        password = request.form['password'].replace(" ", "") 
        
        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verifica que la cédula se encuentre en la tabla data
        cursor.execute('SELECT * FROM data WHERE Cedula = %s', (cedula,))
        data_user = cursor.fetchone()
        
        if not data_user:
            cursor.close()
            return render_template("regisLogin.html", error="Usted no forma parte del personal activo")
        
        # Verifica que la cédula ya esté registrada en la tabla tabla
        cursor.execute('SELECT * FROM tabla WHERE Cedula = %s', (cedula,))
        existing_user = cursor.fetchone()
        
        if existing_user:
            cursor.close()
            return render_template("regisLogin.html", error="La cédula ya está registrada en el sistema")
        
        cursor.execute('SELECT * FROM tabla WHERE username = %s', (username,))
        existing_username = cursor.fetchone()
        
        if existing_username:
            cursor.close()
            return render_template("regisLogin.html", error="El nombre de usuario no está disponible")
        
        # Hash de la contraseña antes de guardarla
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

        # Registrar el nuevo usuario en la tabla tabla
        cursor.execute('INSERT INTO tabla (Cedula, username, Password) VALUES (%s, %s, %s)', 
                       (cedula, username, hashed_password))
        MySQL.connection.commit()
  
        cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', 
                       (cedula, username, 'Registro al usuario {0}'.format(cedula), datetime.now()))
        MySQL.connection.commit()
        cursor.close()
        
        return render_template("regisLogin.html", success="Registro exitoso.")
    
    return render_template("regisLogin.html")

@app.route("/", methods=["GET", "POST"])
def consult():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    super_admin = session.get('Super_Admin')
    fecha = request.form.get('fecha', None)
    tipo_usuario = request.form.get('tipo_usuario', 'general')
    cedula = request.form.get('cedula', None)
    session['cedula_titular'] = cedula
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    print("Cédula recibida desde la sesión:",session['cedula_titular'])
    if super_admin == 0:
        # Si el usuario es básico, solo puede ver los datos del día actual
        fecha = datetime.now().strftime('%Y-%m-%d')
    
    if cedula:
        cursor.execute('''
            SELECT data.Cedula, data.Name_Com, data.Location_Physical, data.Type, data.Location_Admin, 
                   IFNULL(delivery.Entregado, 0) AS Entregado, delivery.Observation, data.ESTADOS, data.Estatus
            FROM data
            LEFT JOIN delivery ON data.ID = delivery.Data_ID
            WHERE data.Cedula = %s
        ''', (cedula,))
        data_exit = cursor.fetchone()
        cursor.close()

        if data_exit:
            estatus = data_exit['Estatus']
            if estatus == 3:
                mensaje = "Suspendido por trámites administrativos."
                return render_template('index.html', super_admin=super_admin, mensaje=mensaje, cedula=cedula)
            elif estatus == 4:
                mensaje = "Suspendido por verificar."
                return render_template('index.html', super_admin=super_admin, mensaje=mensaje, cedula=cedula)
            elif estatus == 2:
                cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus = 2')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus = 2')
                total_recibido = cursor.fetchone()['total_recibido']
                # total_alert = int(total_recibido)
                # alert = None
                # alert_limite = None
                # if total_alert == 550:
                #     alert = f"{total_alert} personas despachadas"
                #     alert_limite = "Se acerca al limite de 600 "
                # elif total_alert == 590:
                #     alert = f"{total_alert} personas despachadas"
                #     alert_limite = "Se acerca al limite de 600"
                # elif total_alert == 600:
                #     alert = f"{total_alert} personas despachadas"
                #     alert_limite = "Ha alcanzado limite de 600"
                faltan = total_personas - total_recibido
                cursor.close()
                return render_template('index.html', super_admin=super_admin, data=data_exit, total_personas=total_personas, total_recibido=total_recibido, faltan=faltan)
            elif estatus == 6:
                mensaje = "Personal Fallecido"
                return render_template('index.html', super_admin=super_admin, mensaje=mensaje, cedula=cedula)
            elif estatus == 5:
                mensaje = "No puede retirar. Está fuera del país."
                return render_template('index.html', super_admin=super_admin, mensaje=mensaje, cedula=cedula)
            elif estatus == 1:
                cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus = 1')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus = 1')
                total_recibido = cursor.fetchone()['total_recibido']
                total_alert = int(total_recibido)
                # alert = None
                # alert_limite = None
                # if total_alert == 550:
                #     alert = f"{total_alert} personas despachadas"
                #     alert_limite = "Se acerca al limite de 600 "
                # elif total_alert == 590:
                #     alert = f"{total_alert} personas despachadas"
                #     alert_limite = "Se acerca al limite de 600"
                # elif total_alert == 600:
                #     alert = f"{total_alert} personas despachadas"
                #     alert_limite = "Ha alcanzado limite de 600"
                faltan = total_personas - total_recibido
                cursor.close()
                return render_template('index.html',  super_admin=super_admin, data=data_exit, total_personas=total_personas, total_recibido=total_recibido, faltan=faltan)
            elif estatus == 10:
                cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus = 1')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus = 1')
                total_recibido = cursor.fetchone()['total_recibido']
                total_alert = int(total_recibido)
                # alert = None
                # alert_limite = None
                # if total_alert == 550:
                #     alert = f"{total_alert} personas despachadas"
                #     alert_limite = "Se acerca al limite de 600 "
                # elif total_alert == 590:
                #     alert = f"{total_alert} personas despachadas"
                #     alert_limite = "Se acerca al limite de 600"
                # elif total_alert == 600:
                #     alert = f"{total_alert} personas despachadas"
                #     alert_limite = "Ha alcanzado limite de 600"
                faltan = total_personas - total_recibido
                cursor.close()
                return render_template('index.html',  super_admin=super_admin, data=data_exit, total_personas=total_personas, total_recibido=total_recibido, faltan=faltan)
            
            elif estatus == 11:
                cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus = 1')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus = 1')
                total_recibido = cursor.fetchone()['total_recibido']
                faltan = total_personas - total_recibido
                cursor.close()
                return render_template('index.html',  super_admin=super_admin, data=data_exit, total_personas=total_personas, total_recibido=total_recibido, faltan=faltan)
            
            elif estatus == 9:
                 mensaje = "Comisión vencida"
                 mensaje2 = 'Comunicarse con el Supervisor o administrador'
                 return render_template('index.html', super_admin=super_admin, mensaje=mensaje, mensaje2=mensaje2,cedula=cedula, mostrar_boton=True)
            
            else:
                mensaje = "Estatus no permitido para retirar"
                mensaje2 = 'Comunicarse con el administrador'
                return render_template('index.html', super_admin=super_admin, mensaje2=mensaje2, mensaje=mensaje, cedula=cedula)
        else:
            cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute('SELECT COUNT(*) AS total_personas FROM data')
            total_personas = cursor.fetchone()['total_personas']
            cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE Entregado = 1')
            total_recibido = cursor.fetchone()['total_recibido']
            faltan = total_personas - total_recibido
            cursor.close()
            mensaje = "Cédula no encontrada"
            return render_template('index.html', super_admin=super_admin, mensaje=mensaje, cedula=cedula, total_personas=total_personas, total_recibido=total_recibido, faltan=faltan)
    else:
        if fecha:
            if tipo_usuario == 'activos':
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus = 1')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus = 1 AND DATE(d.Time_box) = %s', (fecha,))
            elif tipo_usuario == 'pasivos':
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus = 2')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus = 2 AND DATE(d.Time_box) = %s', (fecha,))
           
            elif tipo_usuario == 'comision_servicios_alert':
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus IN (9, 11)')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus IN (9, 11) ')
           
            elif tipo_usuario == 'comision_servicios_2':
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus = 10')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus = 10')
            else:
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE Entregado = 1 AND DATE(Time_box) = %s', (fecha,))
        else:
            if tipo_usuario == 'activos':
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus = 1')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus = 1')
            elif tipo_usuario == 'pasivos':
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data WHERE Estatus = 2')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery d JOIN data ON d.Data_ID = data.ID WHERE d.Entregado = 1 AND data.Estatus = 2')
            else:
                cursor.execute('SELECT COUNT(*) AS total_personas FROM data')
                total_personas = cursor.fetchone()['total_personas']
                cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE Entregado = 1')
        total_recibido = cursor.fetchone()['total_recibido']
        faltan = total_personas - total_recibido
        cursor.close()
    return render_template('index.html', super_admin=super_admin, total_personas=total_personas, total_recibido=total_recibido, faltan=faltan, fecha=fecha, tipo_usuario=tipo_usuario)


# estatus de comicion vencida
@app.route("/cambiar_estatusComision", methods=["POST"])
def cambiar_estatusComision():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    cedula = request.form.get('cedula')
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # Actualizar el estatus a 11
        cursor.execute('UPDATE data SET Estatus = 11 WHERE Cedula = %s AND Estatus = 9', (cedula,))
        MySQL.connection.commit()
        
        # Registrar el cambio en el historial
        cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', 
                       (session['cedula'], session['username'], f'Cambio de estatus de la cédula {cedula} de 9 a 11', datetime.now()))
        MySQL.connection.commit()
        
        mensaje = "Estatus cambiado exitosamente."
    except Exception as e:
        MySQL.connection.rollback()
        mensaje = f"Error al cambiar el estatus: {str(e)}"
    finally:
        cursor.close()
    
    return redirect(url_for('consult', mensaje=mensaje))

# .................

# entrega de beneficio
@app.route("/registrar", methods=["POST"])
def registrar():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    cedula = request.form['cedula']
  
    cedula_personal = request.form['cedula_personal']
    super_admin = session.get('Super_Admin')
    fecha = request.form.get('fecha', None)
    CIFamily = request.form.get('cedulafamiliar', None)
    lunch = request.form.get('lunch', '0')
    if super_admin == 0:
        # Si el usuario es básico, solo puede ver los datos del día actual
        fecha = datetime.now().strftime('%Y-%m-%d')
        
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    
  # Verificar si la cédula del autorizado o familiar ya existe  
    cursor.execute('''
       SELECT COUNT(*) AS total
       FROM data
        WHERE Cedula_autorizado = %s
    ''', (CIFamily,))
    existe_cedula = cursor.fetchone()['total']

    if existe_cedula > 0:
       cursor.close()
       mensaje = "La cédula del autorizado  ya está registrada en el sistema."
       return render_template('index.html', mensaje=mensaje, cedula=cedula, mensaje2="Por favor, verifique los datos ingresados.")
    
    cursor.execute('SELECT * FROM data WHERE Cedula = %s', (cedula,))
    data_exit = cursor.fetchone()
    
    if not data_exit:
        cursor.close()
        mensaje = "La cédula no se encuentra en la tabla data."
        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT COUNT(*) AS total_personas FROM data')
        total_personas = cursor.fetchone()['total_personas']
        cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE Entregado = 1 AND DATE(Time_box) = %s', (fecha,))
        total_recibido = cursor.fetchone()['total_recibido']
        faltan = total_personas - total_recibido
        cursor.close()
        return render_template('index.html', mensaje=mensaje, cedula=cedula, mensaje2="Por favor, verifique la cédula ingresada.", total_personas=total_personas, total_recibido=total_recibido, faltan=faltan)
    
    # Registro exitoso
    observacion = request.form.get('observacion', '').upper()
    nameFamily = request.form.get('nombrefamiliar', None).upper() if request.form.get('nombrefamiliar', None) else None
    CIFamily = CIFamily.upper() if CIFamily else None
    hora_entrega = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    estatus = data_exit['Estatus']
    if estatus == 1:
    # Insertar en la tabla delivery para estatus 1
        cursor.execute('INSERT INTO delivery (Time_box, Staff_ID, Observation, Name_Family, Cedula_Family, Data_ID, Entregado, Lunch) VALUES (%s, %s, %s, %s, %s, %s, 1, %s)', 
                   (hora_entrega, cedula_personal, observacion, nameFamily, CIFamily, data_exit['ID'], lunch))

    elif estatus == 2:
    # Insertar en la tabla delivery para estatus 2
        cursor.execute('''
        INSERT INTO delivery (Time_box, Staff_ID, Observation, Data_ID, Entregado, Lunch) 
        VALUES (%s, %s, %s, %s, 1, %s)
    ''', (hora_entrega, cedula_personal, observacion, data_exit['ID'], lunch))

    # Verificar si los campos `Cedula_autorizado` y `Nombre_autorizado` están vacíos
        if not data_exit['Cedula_autorizado'] and not data_exit['Nombre_autorizado']:
        # Actualizar los datos en los campos si están vacíos
            cursor.execute('''
            UPDATE data
            SET Cedula_autorizado = %s, Nombre_autorizado = %s 
            WHERE Cedula = %s
        ''', (CIFamily, nameFamily, cedula))
    else:
        # Si los campos ya tienen datos, no hacer nada
        print("Los campos Cedula_autorizado y Nombre_autorizado ya tienen datos. No se realizará ninguna modificación.")

# Confirmar los cambios en la base de datos
    MySQL.connection.commit()
    
    if data_exit['Cedula']:
        cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', (session['cedula'], session['username'], 'Marco Como Entregado el Beneficio Alimenticio de {0}'.format(data_exit['Cedula']), session['time_login']))
    MySQL.connection.commit()
    
    mensaje = "Registro exitoso."
    mensaje2 = "El registro se ha completado correctamente."
    
  
    cursor.execute('SELECT COUNT(*) AS total_personas FROM data')
    total_personas = cursor.fetchone()['total_personas']
    cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE Entregado = 1 AND DATE(Time_box) = %s', (fecha,))
    total_recibido = cursor.fetchone()['total_recibido']
    faltan = total_personas - total_recibido
    cursor.close()
    
    return render_template('index.html', mensaje=mensaje, cedula=cedula, mensaje2=mensaje2, total_personas=total_personas, total_recibido=total_recibido, faltan=faltan)

# 78078
# autorizado

@app.route("/obtener_autorizados", methods=["GET"])
def obtener_autorizados():
    if 'loggedin' not in session:
        return jsonify({"error": "No autorizado"}), 403

    # Recupera la cédula del titular desde la sesión
    cedula_titular = session.get('cedula_titular')
    print("Cédula recibida desde la sesión:", cedula_titular)

    if not cedula_titular:
        return jsonify({"error": "Cédula del titular no proporcionada"}), 400

    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('''
        SELECT Cedula_autorizado, Nombre_autorizado, estatus
        FROM data
        WHERE Cedula = %s AND (Cedula_autorizado IS NOT NULL AND TRIM(Cedula_autorizado) != '')
    ''', (cedula_titular,))
    autorizados = cursor.fetchall()
    cursor.close()

    if not autorizados:
        return jsonify({"error": "No se encontraron autorizados para esta cédula, desea asignarlo?"}), 404

    return jsonify(autorizados)

# tabla de autorizados

@app.route("/listado_autorizados", methods=["GET"])
def listado_autorizados():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Consulta para obtener los datos del personal con autorizados
    cursor.execute('''
        SELECT 
            data.Cedula AS Beneficiario_Cedula,
            data.Name_Com AS Beneficiario_Nombre,
            data.Cedula_autorizado AS Autorizado_Cedula,
            data.Nombre_autorizado AS Autorizado_Nombre,
            delivery.Cedula_Family AS Delivery_Cedula,
            delivery.Name_Family AS Delivery_Nombre
        FROM data
        LEFT JOIN delivery ON data.ID = delivery.Data_ID
        WHERE 
            (data.Cedula_autorizado IS NOT NULL AND TRIM(data.Cedula_autorizado) != '') OR
            (delivery.Cedula_Family IS NOT NULL AND TRIM(delivery.Cedula_Family) != '') OR
            (delivery.Name_Family IS NOT NULL AND TRIM(delivery.Name_Family) != '')
    ''')
    
    registros = cursor.fetchall()
    cursor.close()
    
    # Formatear los datos para la tabla
    for registro in registros:
        registro['Autorizado_Cedula'] = registro['Autorizado_Cedula'] or registro['Delivery_Cedula'] or "N/A"
        registro['Autorizado_Nombre'] = registro['Autorizado_Nombre'] or registro['Delivery_Nombre'] or "N/A"
    
    total_autorizados = len(registros)
    return render_template('listado_autorizados.html', registros=registros,total_autorizados=total_autorizados)


# EDITAR USUARIO

@app.route("/editar_autorizado/<cedula>", methods=["GET", "POST"])
def editar_autorizado(cedula):
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == "POST":
        # Obtener los datos enviados desde el formulario
        nuevo_nombre = request.form['nuevo_nombre'].strip()
        nueva_cedula = request.form['nueva_cedula'].strip()
        
        # Actualizar los datos del autorizado en la tabla `data`
        cursor.execute('''
            UPDATE data
            SET Nombre_autorizado = %s, Cedula_autorizado = %s
            WHERE Cedula = %s
        ''', (nuevo_nombre, nueva_cedula, cedula))
        
        # Confirmar los cambios
        MySQL.connection.commit()
        cursor.close()
        
        # Redirigir de vuelta al listado de autorizados
        return redirect(url_for('listado_autorizados'))
    
    # Obtener los datos actuales del autorizado
    cursor.execute('''
        SELECT 
            data.Cedula AS Beneficiario_Cedula,
            data.Name_Com AS Beneficiario_Nombre,
            data.Cedula_autorizado AS Autorizado_Cedula,
            data.Nombre_autorizado AS Autorizado_Nombre
        FROM data
        WHERE data.Cedula = %s
    ''', (cedula,))
    
    registro = cursor.fetchone()
    cursor.close()
    
    return render_template('editar_autorizado.html', registro=registro)

#CONTEO DE ENTREGAS POR USUARIO

@app.route("/reporte_entregas_usuario", methods=["GET", "POST"])
def reporte_entregas_usuario():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == "POST":
        fecha = request.form['fecha']
        if fecha:
            cursor.execute('''
                SELECT Staff_ID, DATE(Time_box) as fecha, COUNT(*) as total_entregas
                FROM delivery
                WHERE Entregado = 1 AND DATE(Time_box) = %s
                GROUP BY Staff_ID, DATE(Time_box)
            ''', (fecha,))
        else:
            cursor.execute('''
                SELECT Staff_ID, DATE(Time_box) as fecha, COUNT(*) as total_entregas
                FROM delivery
                WHERE Entregado = 1
                GROUP BY Staff_ID, DATE(Time_box)
            ''')
    else:
        cursor.execute('''
            SELECT Staff_ID, DATE(Time_box) as fecha, COUNT(*) as total_entregas
            FROM delivery
            WHERE Entregado = 1
            GROUP BY Staff_ID, DATE(Time_box)
        ''')
    
    reportes = cursor.fetchall()
    cursor.close()
    
    for reporte in reportes:
        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT Name_Com FROM data WHERE Cedula = %s', (reporte['Staff_ID'],))
        staff_data = cursor.fetchone()
        cursor.close()
    
    # Manejar el caso en el que no se encuentre el Staff_ID
        staff_name = staff_data['Name_Com'] if staff_data else "Desconocido"
    
    # Actualizar el diccionario `reporte` con el nombre del staff y la fecha formateada
        reporte['staff_name'] = staff_name
        fecha = reporte['fecha']
        reporte['fecha_formateada'] = format_date(fecha, format='full', locale='es_ES')

    
    return render_template('reporte_entregas_usuario.html', reportes=reportes)

@app.route("/reporte_entregas_usuario_excel", methods=["GET", "POST"])
def reporte_entregas_usuario_excel():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # Obtener la fecha del formulario
    fecha = request.form.get('fecha', None)
    
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Filtrar por fecha si se proporciona
    if fecha:
        cursor.execute('''
            SELECT Staff_ID, DATE(Time_box) as fecha, COUNT(*) as total_entregas
            FROM delivery
            WHERE Entregado = 1 AND DATE(Time_box) = %s
            GROUP BY Staff_ID, DATE(Time_box)
        ''', (fecha,))
    else:
        cursor.execute('''
            SELECT Staff_ID, DATE(Time_box) as fecha, COUNT(*) as total_entregas
            FROM delivery
            WHERE Entregado = 1
            GROUP BY Staff_ID, DATE(Time_box)
        ''')
    
    reportes = cursor.fetchall()
    cursor.close()
    
    for reporte in reportes:
        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT Name_Com FROM data WHERE Cedula = %s', (reporte['Staff_ID'],))
        staff_data = cursor.fetchone()
        cursor.close()
    
    # Manejar el caso en el que no se encuentre el Staff_ID
        staff_name = staff_data['Name_Com'] if staff_data else "Desconocido"
    
    # Actualizar el diccionario `reporte` con el nombre del staff y la fecha formateada
        reporte['staff_name'] = staff_name
        fecha = reporte['fecha']
        reporte['fecha_formateada'] = format_date(fecha, format='full', locale='es_ES')
    
    # Generar el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Entregas Usuario"

    headers = ["Fecha", "Usuario", "Total Entregas"]
    ws.append(headers)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.fill = header_fill

    for reporte in reportes:
        row = [
            reporte['fecha_formateada'],
            reporte['staff_name'],
            reporte['total_entregas']
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    column_widths = {
        'A': 20,
        'B': 30,
        'C': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

  # Obtener la fecha actual para el nombre del archivo
    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    nombre_archivo = f"Reporte de entregas por usuario_{fecha_actual}.xlsx"

   
    return send_file(output,download_name=nombre_archivo, as_attachment=True)

# EDITAR EL ESTATUS DE LOS USUARIOS
@app.route("/cambiar_estatus", methods=["GET", "POST"])
def cambiar_estatus():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == "POST":
        cedula = request.form['cedula']
        nuevo_estatus = request.form['estatus']
        
        
        cursor.execute('SELECT Estatus FROM data WHERE Cedula = %s', (cedula,))
        estatus_actual = cursor.fetchone()['Estatus']
        
       
        estatus_nombres = {
            1: "Activo",
            2: "Pasivo",
            3: "Suspendido",
            4: "Suspendido",
            5: "Fuera del país",
            6: "Por verificar"
        }
        
        estatus_actual_nombre = estatus_nombres.get(estatus_actual, "Desconocido")
        nuevo_estatus_nombre = estatus_nombres.get(int(nuevo_estatus), "Desconocido")
        
       
        cursor.execute('UPDATE data SET Estatus = %s WHERE Cedula = %s', (nuevo_estatus, cedula))
        MySQL.connection.commit()
        

        cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', 
                       (session['cedula'], session['username'], f'Cambió el estatus de la cédula {cedula} de {estatus_actual_nombre} a {nuevo_estatus_nombre}', datetime.now()))
        MySQL.connection.commit()
        
        cursor.close()
        return redirect(url_for('cambiar_estatus'))
    
    cursor.execute('SELECT Cedula, Code, Name_Com, Estatus FROM data ')
    usuarios = cursor.fetchall()
    cursor.execute('SELECT COUNT(*) AS total_personas FROM data')
    total_personas = cursor.fetchone()['total_personas']
    cursor.close()
    return render_template('cambiar_estatus.html', usuarios=usuarios, total_personas=total_personas)

@app.route("/editar_usuario/<int:cedula>", methods=["GET", "POST"])
def editar_usuario(cedula):
   
    
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    if request.method == "POST":
        username = request.form['username']
        password = request.form['password'].replace(" ", "") 
        unidad_fisica = request.form['unidad_fisica']
        
       
        if password:
            hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        else:
            hashed_password = None  
       
        if hashed_password:
            cursor.execute('''
                UPDATE tabla
                SET username = %s, Password = %s
                WHERE Cedula = %s
            ''', (username, hashed_password, cedula))
        else:
            cursor.execute('''
                UPDATE tabla
                SET username = %s
                WHERE Cedula = %s
            ''', (username, cedula))
        
        cursor.execute('''
            UPDATE data
            SET Location_Physical = %s
            WHERE Cedula = %s
        ''', (unidad_fisica, cedula))
        MySQL.connection.commit()
        cursor.close()
        return redirect(url_for('usuarios'))
    
    # Obtener los datos actuales del usuario
    cursor.execute('''
        SELECT tabla.Cedula, data.Name_Com, tabla.username, data.Location_Physical, tabla.Password, tabla.estado
        FROM tabla
        JOIN data ON tabla.Cedula = data.Cedula
        WHERE tabla.Cedula = %s
    ''', (cedula,))
    usuario = cursor.fetchone()
    cursor.close()
    
    return render_template('editar_usuario.html', usuario=usuario)


@app.route("/suspender_usuario/<int:cedula>", methods=["POST"])
def suspender_usuario(cedula):
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('UPDATE tabla SET estado = %s WHERE Cedula = %s', ('suspendido', cedula))
    MySQL.connection.commit()
    cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', (session['cedula'], session['username'], 'Suspendio El Usuario {0}'.format(cedula),session['time_login']))
    MySQL.connection.commit()
    cursor.close()
    return redirect(url_for('usuarios'))

@app.route("/reactivar_usuario/<int:cedula>", methods=["POST"])
def reactivar_usuario(cedula):
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('UPDATE tabla SET estado = %s WHERE Cedula = %s', ('activo', cedula))
    MySQL.connection.commit()
    cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', (session['cedula'], session['username'], 'Reactivo El Usuario {0}'.format(cedula),session['time_login']))
    MySQL.connection.commit()
    cursor.close()
    return redirect(url_for('usuarios'))

@app.route("/usuarios", methods=["GET"])
def usuarios():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('''
        SELECT tabla.Cedula, data.Name_Com, tabla.username, data.Location_Physical, data.Location_Admin, tabla.estado
        FROM tabla
        JOIN data ON tabla.Cedula = data.Cedula
    ''')
    usuarios = cursor.fetchall()
    cursor.close()
    return render_template('usuarios.html', usuarios=usuarios)

@app.route("/reporte_usuarios", methods=["GET","POST"])
def reporte_usuarios():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('select * from user_history')
    historial = cursor.fetchall()
    MySQL.connection.commit()
   
    cursor.close()
    return render_template('reporte_usuarios.html', historial=historial)

# REGISTRAR NUEVOS EMPLEADOS ACTIVOS/PASIVOS
@app.route("/nuevoEmpActivo", methods=["GET", "POST"])
def NuevoUserActivo():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    if request.method == "POST":
        cedula = request.form['cedula']
        nombreCompleto = request.form['nombreCompleto']
        unidadFisica = request.form['unidadFisica']
        unidadAdmin = request.form['unidadAdmin']
        observacion = request.form['observacion']
        CIFamiliar = request.form.get('cedula-family', None)
        Nombre_Familiar = request.form.get('Nombre_Familiar', None)
        CodigoCarnet = request.form.get('CodigoCarnet', None)
        estatus = 1  
        cedula_personal = session['cedula']
        entregado = True if 'entregado' in request.form else False
        lunch = True if 'lunch' in request.form else False
        horaEntrega = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        try:
            cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute('SELECT * FROM data WHERE Cedula = %s', (cedula,))
            existing_user = cursor.fetchone()

            if existing_user:
                cursor.close()
                return render_template("nuevo_user_activo.html", error="Esta cédula ya está registrada.")

           
            cursor.execute('SELECT MAX(ID) AS max_id FROM data')
            max_id = cursor.fetchone()['max_id']
            new_id = max_id + 1 if max_id is not None else 1

            cursor.execute('INSERT INTO data (ID, Cedula, Name_Com, Code, Location_Physical, Location_Admin, manually, Estatus) VALUES (%s, %s, %s, %s, %s, %s, 1, %s)', 
                           (new_id, cedula, nombreCompleto, CodigoCarnet, unidadFisica, unidadAdmin, estatus))
            MySQL.connection.commit()
            cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', 
                           (session['cedula'], session['username'], f'Registró un personal activo con cédula {cedula}', datetime.now()))
            MySQL.connection.commit()

            if entregado:
                cursor.execute('SELECT ID FROM data WHERE Cedula = %s', (cedula,))
                data_id = cursor.fetchone()
                cursor.execute('INSERT INTO delivery (Time_box, Staff_ID, Name_Family, Cedula_Family, Observation, Data_ID, Entregado, Lunch) VALUES (%s, %s, %s, %s, %s, %s, 1, %s)', 
                               (horaEntrega, cedula_personal, Nombre_Familiar, CIFamiliar if CIFamiliar else None, observacion, data_id['ID'], lunch))

                MySQL.connection.commit()
                cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', 
                               (session['cedula'], session['username'], f'Marcó como entregado el beneficio alimenticio para la cédula {cedula}', datetime.now()))
                MySQL.connection.commit()
            cursor.close()
            return render_template("nuevo_user_activo.html", success="Registro exitoso.")  
        except Exception as e:
            return render_template("nuevo_user_activo.html", error=f"Error en el registro: {str(e)}")
    return render_template("nuevo_user_activo.html")

@app.route("/nuevoEmpPasivo", methods=["GET", "POST"])
def NuevoUserPasivo():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    if request.method == "POST":
        cedula = request.form['cedula']
        nombreCompleto = request.form['nombreCompleto']
        observacion = request.form['observacion']
        CIFamiliar = request.form.get('cedula-family', None)
        Nombre_Familiar = request.form.get('Nombre_Familiar', None)
        CodigoCarnet = request.form.get('CodigoCarnet', None)
        estatus = 2  
        estado = request.form['estado']
        cedula_personal = session['cedula']
        entregado = True if 'entregado' in request.form else False
        lunch = True if 'lunch' in request.form else False
        horaEntrega = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        try:
            cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute('SELECT * FROM data WHERE Cedula = %s', (cedula,))
            existing_user = cursor.fetchone()

            if existing_user:
                cursor.close()
                return render_template("nuevo_user_pasivo.html", error="Esta cédula ya está registrada.")

            cursor.execute('SELECT MAX(ID) AS max_id FROM data')
            max_id = cursor.fetchone()['max_id']
            new_id = max_id + 1 if max_id is not None else 1

            cursor.execute('INSERT INTO data (ID, Cedula, Name_Com, Code, manually, Estatus, ESTADOS) VALUES (%s, %s, %s, %s, 1, %s, %s)', 
                           (new_id, cedula, nombreCompleto, CodigoCarnet, estatus, estado))
            MySQL.connection.commit()
            cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', 
                           (session['cedula'], session['username'], f'Registró un personal pasivo con cédula {cedula}', datetime.now()))
            MySQL.connection.commit()

            if entregado:
                cursor.execute('SELECT ID FROM data WHERE Cedula = %s', (cedula,))
                data_id = cursor.fetchone()
                cursor.execute('INSERT INTO delivery (Time_box, Staff_ID, Name_Family, Cedula_Family, Observation, Data_ID, Entregado, Lunch) VALUES (%s, %s, %s, %s, %s, %s, 1, %s)', 
                               (horaEntrega, cedula_personal, Nombre_Familiar, CIFamiliar if CIFamiliar else None, observacion, data_id['ID'], lunch))

                MySQL.connection.commit()
                cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', 
                               (session['cedula'], session['username'], f'Marcó como entregado el beneficio alimenticio para la cédula {cedula}', datetime.now()))
                MySQL.connection.commit()
            cursor.close()
            return render_template("nuevo_user_pasivo.html", success="Registro exitoso.")  
        except Exception as e:
            return render_template("nuevo_user_pasivo.html", error=f"Error en el registro: {str(e)}")
    return render_template("nuevo_user_pasivo.html")

# MOSTRAR REPORTE DE ENTREGADOS
@app.route("/reporte")
def reporte():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('''
        SELECT DATE(Time_box) as fecha,
               SUM(CASE WHEN data.Estatus = 1 THEN 1 ELSE 0 END) as total_activos,
               SUM(CASE WHEN data.Estatus = 2 THEN 1 ELSE 0 END) as total_pasivos,
               SUM(CASE WHEN data.Estatus  IN (9, 11)  THEN 1 ELSE 0 END) as total_comision_vencida,
               SUM(CASE WHEN data.Estatus = 10 THEN 1 ELSE 0 END) as total_comision_vigente,
               COUNT(*) as total_entregas
        FROM delivery
        JOIN data ON delivery.Data_ID = data.ID
        WHERE Entregado = 1
        GROUP BY DATE(Time_box)
    ''')
    reportes = cursor.fetchall()

    # Calcular totales generales
    total_entregas = sum(reporte['total_entregas'] for reporte in reportes)
    total_activos = sum(reporte['total_activos'] for reporte in reportes)
    total_pasivos = sum(reporte['total_pasivos'] for reporte in reportes)
    total_comision_vencida = sum(reporte['total_comision_vencida'] for reporte in reportes)
    total_comision_vigente = sum(reporte['total_comision_vigente'] for reporte in reportes)

    cursor.close()

    # Formatear las fechas
    for reporte in reportes:
        fecha = reporte['fecha']
        reporte['fecha_formateada'] = format_date(fecha, format='full', locale='es_ES')

    return render_template(
        'reporte.html',
        reportes=reportes,
        total_entregas=total_entregas,
        total_activos=total_activos,
        total_pasivos=total_pasivos,
        total_comision_vencida=total_comision_vencida,
        total_comision_vigente=total_comision_vigente
    )
# GENERAR REPORTE DE ENTREGADOS EN PDF
@app.route("/reporte_pdf", methods=["GET", "POST"])
def reporte_pdf():
    if request.method == "POST":
        fecha = request.form['fecha']
        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT DATE(Time_box) as fecha,
                   SUM(CASE WHEN data.Estatus = 1 THEN 1 ELSE 0 END) as total_activos,
                   SUM(CASE WHEN data.Estatus = 2 THEN 1 ELSE 0 END) as total_pasivos,
                   SUM(CASE WHEN data.Estatus IN (9, 11) THEN 1 ELSE 0 END) as total_comision_vencida,
                   SUM(CASE WHEN data.Estatus = 10 THEN 1 ELSE 0 END) as total_comision_vigente,
                   COUNT(*) as total_entregas
            FROM delivery
            JOIN data ON delivery.Data_ID = data.ID
            WHERE Entregado = 1 AND DATE(Time_box) = %s
            GROUP BY DATE(Time_box)
        ''', (fecha,))
    else:
        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT DATE(Time_box) as fecha,
                   SUM(CASE WHEN data.Estatus = 1 THEN 1 ELSE 0 END) as total_activos,
                   SUM(CASE WHEN data.Estatus = 2 THEN 1 ELSE 0 END) as total_pasivos,
                   SUM(CASE WHEN data.Estatus IN (9, 11) THEN 1 ELSE 0 END) as total_comision_vencida,
                   SUM(CASE WHEN data.Estatus = 10 THEN 1 ELSE 0 END) as total_comision_vigente,
                   COUNT(*) as total_entregas
            FROM delivery
            JOIN data ON delivery.Data_ID = data.ID
            WHERE Entregado = 1
            GROUP BY DATE(Time_box)
        ''')
    reportes = cursor.fetchall()
    cursor.close()
    
    # Calcular totales generales
    total_entregas = sum(reporte['total_entregas'] for reporte in reportes)
    total_activos = sum(reporte['total_activos'] for reporte in reportes)
    total_pasivos = sum(reporte['total_pasivos'] for reporte in reportes)
    total_comision_vencida = sum(reporte['total_comision_vencida'] for reporte in reportes)
    total_comision_vigente = sum(reporte['total_comision_vigente'] for reporte in reportes)
    
    # Formatear las fechas
    for reporte in reportes:
        fecha = reporte['fecha']
        reporte['fecha_formateada'] = format_date(fecha, format='full', locale='es_ES')
    
    rendered = render_template(
        'reporte_pdf.html',
        reportes=reportes,
        total_entregas=total_entregas,
        total_activos=total_activos,
        total_pasivos=total_pasivos,
        total_comision_vencida=total_comision_vencida,
        total_comision_vigente=total_comision_vigente
    )
    pdf = HTML(string=rendered).write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=reporte.pdf'
    return response

# MOSTRAR LISTADO DE ENTREGADOS 
@app.route("/listado")
def listado():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute('''
    SELECT d.Data_ID, d.Time_box, d.Entregado, d.Staff_ID, d.Observation, d.Lunch, d.Cedula_Family, d.Name_Family,data.Cedula_autorizado, data.Nombre_autorizado,
           data.Cedula, data.Code, data.Name_Com, data.Manually, data.Location_Admin, data.Estatus, data.ESTADOS, data.Location_Physical,
           staff.Name_Com AS Registrador_Name  -- Incluye el nombre del registrador
    FROM delivery d
    JOIN data ON d.Data_ID = data.ID
    LEFT JOIN data AS staff ON d.Staff_ID = staff.Cedula  -- Une con la tabla de datos para obtener el nombre del registrador
    WHERE d.Entregado = 1
    ''')
    registros = cursor.fetchall()
       
    cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE Entregado = 1')
    total_recibido = cursor.fetchone()['total_recibido']
    cursor.close()
    return render_template('tabla.html', registros=registros, total_recibido=total_recibido)

# GENERAR LISTADO DE ENTREGADOS EN PDF
@app.route("/listado_pdf", methods=["GET", "POST"])
def listado_pdf():
    if request.method == "POST":
        fecha = request.form['fecha']
        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT d.Data_ID, d.Time_box, d.Entregado,  d.Staff_ID, d.Observation, d.Lunch,d.Cedula_Family, d.Name_Family,data.Cedula_autorizado, data.Nombre_autorizado,
                   data.Cedula, data.Name_Com,data.Manually,data.Location_Admin, data.Estatus,data.ESTADOS,data.Location_Physical 
            FROM delivery d
            JOIN data ON d.Data_ID = data.ID
            WHERE d.Entregado = 1 AND DATE(d.Time_box) = %s
        ''', (fecha,))
        registros = cursor.fetchall()
        cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE Entregado = 1 AND DATE(Time_box) = %s', (fecha,))
        total_recibido = cursor.fetchone()['total_recibido']
        cursor.close()

        rendered = render_template('tabla_pdf.html', registros=registros, total_recibido=total_recibido, fecha=fecha)
        pdf = HTML(string=rendered).write_pdf()

        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=listado.pdf'
        return response
    else:
        return render_template('tabla_pdf.html')
    
# GENERAR LISTADO EN EXCEL
@app.route("/listado_excel", methods=["GET", "POST"])
def listado_excel():
    if request.method == "POST":
        fecha = request.form['fecha']
        filtro = request.form.get('filtro', 'todos')  # Obtiene el filtro seleccionado

        # Construir la consulta SQL dinámicamente
        query = '''
            SELECT d.Data_ID, d.Time_box, d.Entregado, d.Staff_ID, d.Observation, d.Lunch,d.Cedula_Family, d.Name_Family,data.Cedula_autorizado, data.Nombre_autorizado,
                   data.Cedula, data.Name_Com, data.Manually, data.Location_Admin, data.Estatus, data.ESTADOS, data.Location_Physical 
            FROM delivery d
            JOIN data ON d.Data_ID = data.ID
            WHERE d.Entregado = 1 AND DATE(d.Time_box) = %s
        '''
        if filtro == 'autorizados':
            query += ' AND (d.Cedula_Family IS NOT NULL OR data.Cedula_autorizado IS NOT NULL)'
        elif filtro == 'activo':
            query += ' AND data.Estatus = 1'
        elif filtro == 'pasivo':
            query += ' AND data.Estatus = 2'
        elif filtro == 'manually':
            query += ' AND data.Manually = 1'
        elif filtro == 'comision_vencida':
            query += ' AND data.Estatus IN (9, 11)'
        elif filtro == 'comision_vigente':
            query += ' AND data.Estatus = 10'

        query += ' ORDER BY data.ESTADOS ASC, data.Location_Physical ASC, data.Location_Admin ASC'

        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(query, (fecha,))
        registros = cursor.fetchall()
        cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE Entregado = 1 AND DATE(Time_box) = %s', (fecha,))
        total_recibido = cursor.fetchone()['total_recibido']
        cursor.close()

        # Generar el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Listado"

        img1_path = os.path.join(app.root_path, 'static/css/img/logo.png')
        img2_path = os.path.join(app.root_path, 'static/css/img/logo2.png')
        img1 = Image(img1_path)
        img2 = Image(img2_path)
        img1.width, img1.height = 60, 60
        img2.width, img2.height = 60, 60
        ws.add_image(img1, 'A1')

        headers = ["#", "Cedula", "Nombre Completo", "Estado", "Estatus", "Unidad Fisica", "Ubicación administrativa", "Hora de Entrega", "Cedula del Autorizado", "Nombre del Autorizado", "Observacion", "Registro Manual", "Merienda", "Cedula del Registrador", "Nombre del Registrador"]

        last_column = chr(64 + len(headers))
        img2.anchor = f'{last_column}1'
        ws.add_image(img2)

        ws.row_dimensions[1].height = 55

        ws.merge_cells(f'A1:{last_column}1')
        ws['A1'] = "Listado de Personas que han Recibido la Caja"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.fill = header_fill

        ws.row_dimensions[2].height = 30

        for idx, registro in enumerate(registros, start=1):
            cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute('SELECT Name_Com FROM data WHERE Cedula = %s', (registro['Staff_ID'],))
            staff_data = cursor.fetchone()
            cursor.close()

            # Manejar el caso en el que no se encuentre el Staff_ID
            staff_name = staff_data['Name_Com'] if staff_data else "Desconocido"
            
            cedula_autorizado = registro['Cedula_Family'] if registro['Cedula_Family'] else registro['Cedula_autorizado'] if registro['Cedula_autorizado'] else ''
            nombre_autorizado = registro['Name_Family'] if registro['Name_Family'] else registro['Nombre_autorizado'] if registro['Nombre_autorizado'] else ''

            estatus = "Activo" if registro['Estatus'] == 1 else "Pasivo" if registro['Estatus'] == 2 else "Comisión Vencida" if registro['Estatus'] == 9 else "Comisión Vencida" if registro['Estatus'] == 11 else "Comisión Vigente"
           
            row = [
                idx,
                registro['Cedula'],
                registro['Name_Com'],
                registro['ESTADOS'] if registro['Estatus'] in [2, 9, 10] else "",
                estatus,
                registro['Location_Physical'] if  registro['Estatus'] == 1 else "",
                registro['Location_Admin'] if   registro['Estatus'] == 1 else "",
                registro['Time_box'],
                cedula_autorizado,
                nombre_autorizado,
                registro['Observation'] if registro['Observation'] else ' ',
                "Si" if registro['Manually'] else 'No',
                "Si" if registro['Lunch'] else 'No',
                registro['Staff_ID'],
                staff_name
            ]
            
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        column_widths = {
            'A': 7,
            'B': 10,
            'C': 20,
            'D': 20,
            'E': 20,
            'F': 20,
            'G': 20,
            'H': 20,
            'I': 20,
            'J': 20,
            'K': 20,
            'L': 20,
            'M': 20,
            'N': 20,
            'O': 20,
            'P': 20
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 82.5

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        nombre_archivo = f"listado_entregados_{fecha_actual}.xlsx"
        
        return send_file(output, download_name=nombre_archivo, as_attachment=True)
    return render_template('tabla_pdf.html')


# GENERAR LISTADO DE NO ENTREGADOS 
@app.route("/listado_no_registrado")
def listado_no_registrado():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('''
        SELECT data.Cedula, data.Name_Com, data.Location_Physical, data.Estatus,data.ESTADOS,data.Code,data.Location_Admin, delivery.Entregado
        FROM data
        LEFT JOIN delivery ON data.ID = delivery.Data_ID
        WHERE delivery.Entregado IS NULL OR delivery.Entregado = 0
    ''')
    registros = cursor.fetchall()
    cursor.execute('''
        SELECT COUNT(*) AS total_no_entregados
        FROM data
        LEFT JOIN delivery ON data.ID = delivery.Data_ID
        WHERE delivery.Entregado IS NULL OR delivery.Entregado = 0
    ''')
    total_no_entregados = cursor.fetchone()['total_no_entregados']
    cursor.close()
    return render_template('listado_no_registrado.html', registros=registros, total_no_entregados=total_no_entregados)

# GENERAR LISTADO DE NO ENTREGADOS EN PDF
@app.route("/listado_no_regist_pdf", methods=["GET", "POST"])
def listado_no_regist_pdf():
    filtro = request.args.get('filtro', 'todos')  # Obtiene el filtro desde la URL (por defecto "todos")
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Construir la consulta SQL dinámicamente según el filtro
    query = '''
        SELECT data.Cedula, data.Name_Com, data.Location_Physical, data.Location_Admin, data.Code, data.Estatus, data.ESTADOS, delivery.Entregado
        FROM data
        LEFT JOIN delivery ON data.ID = delivery.Data_ID
        WHERE delivery.Entregado IS NULL OR delivery.Entregado = 0
    '''
    if filtro == 'activos':
        query += ' AND data.Estatus = 1'
    elif filtro == 'pasivos':
        query += ' AND data.Estatus = 2'
    elif filtro == 'comision_vigente':
        query += ' AND data.Estatus = 10'
    elif filtro == 'comision_vencida':
        query += ' AND data.Estatus IN (9, 11)'
    
    cursor.execute(query)
    registros = cursor.fetchall()
    
    # Contar el total de registros según el filtro
    cursor.execute(f'''
        SELECT COUNT(*) AS total_no_entregados
        FROM data
        LEFT JOIN delivery ON data.ID = delivery.Data_ID
        WHERE delivery.Entregado IS NULL OR delivery.Entregado = 0
        {'AND data.Estatus = 1' if filtro == 'activos' else ''}
        {'AND data.Estatus = 2' if filtro == 'pasivos' else ''}
        {'AND data.Estatus = 10' if filtro == 'comision_vigente' else ''}
        {'AND data.Estatus IN (9, 11)' if filtro == 'comision_vencida' else ''}
    ''')
    total_no_entregados = cursor.fetchone()['total_no_entregados']
    cursor.close()
    
    rendered = render_template('tabla_no_regist_pdf.html', registros=registros, total_no_entregados=total_no_entregados, filtro=filtro)
    pdf = HTML(string=rendered).write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=listado.pdf'
    return response

# Generar listado de no entregado en excel
@app.route("/listado_no_registrado_excel", methods=["GET", "POST"])
def listado_no_registrado_excel():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # Obtener el filtro de estatus desde el formulario
    filtro = request.form.get('filtro', 'todos')  # Por defecto, "todos"
    print(f"Filtro seleccionado: {filtro}")

    # Construir la consulta SQL dinámicamente según el filtro
    query = '''
        SELECT data.Cedula, data.Name_Com, data.Location_Physical, data.Location_Admin, data.Code, data.Estatus, data.ESTADOS, delivery.Entregado
        FROM data
        LEFT JOIN delivery ON data.ID = delivery.Data_ID
        WHERE delivery.Entregado IS NULL OR delivery.Entregado = 0
    '''
    if filtro == 'activo':
        query += ' AND data.Estatus = 1'
    elif filtro == 'pasivo':
        query += ' AND data.Estatus = 2'
    elif filtro == 'comision_vigente':
        query += ' AND data.Estatus = 10'
    elif filtro == 'comision_vencida':
        query += ' AND data.Estatus IN (9, 11)'
    elif filtro == 'autorizados':
        query += ' AND (data.Cedula_autorizado IS NOT NULL OR TRIM(data.Cedula_autorizado) != "")'

    # Ejecutar la consulta
    cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute(query)
    registros = cursor.fetchall()
    cursor.close()

    # Generar el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "No Entregados"

    # Agregar imágenes
    img1_path = os.path.join(app.root_path, 'static/css/img/logo.png')
    img2_path = os.path.join(app.root_path, 'static/css/img/logo2.png')
    img1 = Image(img1_path)
    img2 = Image(img2_path)
    img1.width, img1.height = 60, 60
    img2.width, img2.height = 60, 60
    ws.add_image(img1, 'A1')

    # Encabezados
    headers = ["#", "Cédula", "Nombre Completo", "Unidad Física", "Ubicación Administrativa", "Código", "Estatus", "Estado", "Entregado"]
    last_column = chr(64 + len(headers))
    img2.anchor = f'{last_column}1'
    ws.add_image(img2)

    ws.row_dimensions[1].height = 55
    ws.merge_cells(f'A1:{last_column}1')
    ws['A1'] = "Listado de Personas que No han Recibido la Caja"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Agregar encabezados
    ws.append(headers)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.fill = header_fill

    ws.row_dimensions[2].height = 30

    # Agregar los datos
    for idx, registro in enumerate(registros, start=1):
        estatus = "Activo" if registro['Estatus'] == 1 else "Pasivo" if registro['Estatus'] == 2 else "Comisión Vencida" if registro['Estatus'] in [9, 11] else "Comisión Vigente" if registro['Estatus'] == 10 else "Desconocido"
        row = [
            idx,
            registro['Cedula'],
            registro['Name_Com'],
            registro['Location_Physical'],
            registro['Location_Admin'],
            registro['Code'],
            estatus,
            registro['ESTADOS'],
            "No" if registro['Entregado'] == 0 or registro['Entregado'] is None else "Sí"
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Ajustar el ancho de las columnas
    column_widths = {
        'A': 7,
        'B': 15,
        'C': 25,
        'D': 20,
        'E': 25,
        'F': 15,
        'G': 20,
        'H': 15,
        'I': 10
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Guardar el archivo en memoria
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Obtener la fecha actual para el nombre del archivo
    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    nombre_archivo = f"listado_no_entregados_{fecha_actual}.xlsx"

    return send_file(output, download_name=nombre_archivo, as_attachment=True)



# ................


# insercion a la DB

def cargar_excel_a_db(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        # Verificar si la solicitud es POST
        if request.method == "POST":
            if 'loggedin' not in session:
                return redirect(url_for('login'))
            
            # Verificar si el usuario tiene permisos de super_admin
            if session.get('Super_Admin') != 1:
                return render_template("gestionar_data.html", error="No tiene permisos para realizar esta acción.")
            if 'file' not in request.files or request.files['file'].filename == '':
                return render_template("gestionar_data.html", error="Debe seleccionar un archivo Excel para cargar datos.")
            
            file = request.files['file']
            if not file.filename.endswith('.xlsx'):
                return render_template("gestionar_data.html", error="El archivo debe ser un archivo Excel (.xlsx).")
            
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.root_path, "uploads", filename)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            file.save(file_path)

            try:
                # Leer el archivo Excel
                df = pd.read_excel(file_path)

                # Verificar que el archivo contiene las columnas necesarias
                required_columns = [
                    "Type", "Cedula", "Name_Com", "Code", "Location_Physical",
                    "Location_Admin", "manually", "suspended", "Estatus", "ESTADOS",
                    "Cedula_autorizado", "Nombre_autorizado"
                ]
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    return render_template("gestionar_data.html", error=f"El archivo Excel no contiene las columnas necesarias: {', '.join(missing_columns)}")
                
                # Reemplazar NaN con valores predeterminados o NULL
                try:
                    # Convertir los tipos de datos primero para evitar problemas con fillna
                    df["Cedula"] = pd.to_numeric(df["Cedula"], errors='coerce')
                    df["Code"] = pd.to_numeric(df["Code"], errors='coerce')
                    df["manually"] = pd.to_numeric(df["manually"], errors='coerce')
                    df["suspended"] = pd.to_numeric(df["suspended"], errors='coerce')
                    df["Estatus"] = pd.to_numeric(df["Estatus"], errors='coerce')
                    
                    text_columns = ["Type", "Name_Com", "Location_Physical", "Location_Admin", 
                                    "ESTADOS", "Cedula_autorizado", "Nombre_autorizado"]
                    for col in text_columns:
                        df[col] = df[col].astype(str).replace('nan', '')
                    
                    df = df.fillna({
                        "Type": "", "Cedula": 0, "Name_Com": "", "Code": 0,
                        "Location_Physical": "", "Location_Admin": "",
                        "manually": 0, "suspended": 0, "Estatus": 1,
                        "ESTADOS": "", "Cedula_autorizado": "", "Nombre_autorizado": ""
                    })
                    
                    int_columns = ["Cedula", "Code", "manually", "suspended", "Estatus"]
                    for col in int_columns:
                        df[col] = df[col].fillna(0).astype(int)
                except Exception as e:
                    return render_template("gestionar_data.html", error=f"Error al procesar los tipos de datos: {str(e)}")

                # Insertar datos en la base de datos con validación de duplicados
                cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
                registros_insertados = 0  # Contador de registros insertados
                for _, row in df.iterrows():
                    # Verificar si la cédula ya existe en la base de datos
                    cursor.execute('SELECT COUNT(*) AS count FROM data WHERE Cedula = %s', (row['Cedula'],))
                    exists = cursor.fetchone()['count']
                    if exists == 0:
                        cursor.execute('''
                            INSERT INTO data (Type, Cedula, Name_Com, Code, Location_Physical, Location_Admin, manually, suspended, Estatus, ESTADOS, Cedula_autorizado, Nombre_autorizado)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ''', (
                            row['Type'], row['Cedula'], row['Name_Com'], row['Code'],
                            row['Location_Physical'], row['Location_Admin'], row['manually'],
                            row['suspended'], row['Estatus'], row['ESTADOS'],
                            row['Cedula_autorizado'], row['Nombre_autorizado']
                        ))
                        registros_insertados += 1
                MySQL.connection.commit()
                cursor.close()
                return func(*args, **kwargs, success=f"Los datos han sido cargados correctamente. Registros insertados: {registros_insertados}.")
            except Exception as e:
                MySQL.connection.rollback()
                return render_template("gestionar_data.html", error=f"Error al procesar el archivo: {str(e)}")
        
        # Si el método no es POST, simplemente renderizar la vista
        return func(*args, **kwargs)
    return wrapper

@app.route("/cargar_data", methods=["GET", "POST"])
@cargar_excel_a_db
def cargar_data(success=None):
    
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # Verificar si el usuario tiene permisos de super_admin
    if session.get('Super_Admin') != 1:
        return render_template("gestionar_data.html", error="No tiene permisos para realizar esta acción.")
    if request.method == "GET":
        # Redirigir a la página principal si se accede con GET
        return redirect(url_for('gestionar_data'))
    return render_template("gestionar_data.html", success=success)


# vaciar la DB
@app.route("/vaciar_db", methods=["POST"])
def vaciar_db():
    # Verificar que el usuario esté autenticado
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # Verificar si el usuario tiene permisos de super_admin
    if session.get('Super_Admin') != 1:
        return render_template("gestionar_data.html", error="No tiene permisos para realizar esta acción.")
    
    cursor = None
    try:
        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verificar conexión a la base de datos
        cursor.execute("SELECT 1")
        if not cursor.fetchone():
            return render_template("gestionar_data.html", error="No se pudo conectar a la base de datos.")
        
        # Iniciar transacción
        cursor.execute("START TRANSACTION")
        
        # Obtener contador antes de vaciar para el registro
        cursor.execute("SELECT COUNT(*) AS total FROM data")
        result = cursor.fetchone()
        total_registros = result['total'] if result else 0
        
        # Verificar que las tablas existen
        cursor.execute("SHOW TABLES LIKE 'delivery'")
        delivery_exists = cursor.fetchone() is not None
        cursor.execute("SHOW TABLES LIKE 'data'")
        data_exists = cursor.fetchone() is not None
        
        if not delivery_exists or not data_exists:
            raise Exception("Las tablas necesarias no existen en la base de datos")
        
        # Desactivar temporalmente verificación de claves foráneas
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")

        # Eliminar primero los registros de delivery ya que tiene FK a data
        cursor.execute("DELETE FROM delivery")
        deleted_delivery = cursor.rowcount
        
        # Eliminar todos los registros de la tabla data
        cursor.execute("DELETE FROM data")
        deleted_data = cursor.rowcount
        
        # Reactivar verificación de claves foráneas
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        
        # Confirmar la transacción
        MySQL.connection.commit()
        
        # Registrar la acción en el historial
        action_message = f'Vació la base de datos. Se eliminaron {total_registros} registros de data y {deleted_delivery} registros de delivery'
        cursor.execute('INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)', 
                      (session['cedula'], session['username'], action_message, datetime.now()))
        MySQL.connection.commit()
        
        success_message = f"Base de datos vaciada correctamente. Se eliminaron {total_registros} registros de data y {deleted_delivery} registros de delivery."
        return render_template("gestionar_data.html", success=success_message)
    
    except Exception as e:
        # Revertir cambios en caso de error
        if cursor:
            try:
                # Asegurar que la verificación de claves foráneas se reactiva incluso en caso de error
                cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
                MySQL.connection.rollback()
            except:
                pass  # Evitar errores en cascada si la conexión ya está corrupta
        
        error_message = f"Error al vaciar la base de datos: {str(e)}"
        return render_template("gestionar_data.html", error=error_message)
    
    finally:
        # Asegurar que el cursor se cierra incluso si hay una excepción
        if cursor:
            try:
                cursor.close()
            except:
                pass  # Ignorar errores al cerrar el cursor


# backud

def backup_to_excel():
    """Genera un backup de la base de datos en formato Excel."""
    # Verificar autenticación y permisos primero
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    if session.get('Super_Admin') != 1:
        return render_template("gestionar_data.html", error="No tiene permisos para realizar esta acción.")
    
    try:
        # Crear directorio de backups si no existe
        backup_dir = os.path.join(app.root_path, "backups")
        os.makedirs(backup_dir, exist_ok=True)

        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'copia_db_{timestamp}.xlsx'
        file_path = os.path.join(backup_dir, filename)

        # Crear un Excel writer para múltiples hojas
        writer = pd.ExcelWriter(file_path, engine='openpyxl')

        # Obtener datos de cada tabla
        cursor = MySQL.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Lista explícita de todas las tablas que deben exportarse, asegurando que 'delivery' esté incluida
        tables = ['data', 'delivery', 'user_history', 'tabla']
        
        # Validar que todas las tablas existen para notificar cualquier problema
        cursor.execute("SHOW TABLES")
        all_db_tables = [row[f"Tables_in_{app.config['MYSQL_DB']}"] for row in cursor.fetchall()]
        tables_exported = []
        
        # Create a log file for Excel backup
        excel_log_path = os.path.join(backup_dir, f"excel_backup_log_{timestamp}.txt")
        with open(excel_log_path, 'w', encoding='utf-8') as excel_log:
            excel_log.write(f"Iniciando backup Excel: {datetime.now()}\n")
            excel_log.write(f"Tablas encontradas en la base de datos: {', '.join(all_db_tables)}\n")
            excel_log.write(f"Tablas a exportar: {', '.join(tables)}\n")
            
            # Verificar si 'delivery' está en la base de datos
            if 'delivery' in all_db_tables:
                excel_log.write("Tabla 'delivery' encontrada en la base de datos - se intentará exportar\n")
            else:
                excel_log.write("ADVERTENCIA: Tabla 'delivery' NO encontrada en la base de datos\n")
            
            for table in tables:
                try:
                    excel_log.write(f"Exportando tabla: {table}\n")
                    # Use parametrized query to avoid SQL injection
                    # Comprobar explícitamente si la tabla existe
                    cursor.execute(f'SHOW TABLES LIKE %s', (table,))
                    table_exists = cursor.fetchone() is not None
                    
                    if table_exists:
                        # Explicitly use parametrized SELECT to ensure proper execution
                        # Usar consulta con manejo de errores explícito
                        try:
                            excel_log.write(f"Ejecutando SELECT en tabla: {table}\n")
                            cursor.execute(f'SELECT * FROM `{table}`')
                            data = cursor.fetchall()
                            excel_log.write(f"Obtenidos {len(data) if data else 0} registros de {table}\n")
                            
                            if data:
                                # Asegurar que los datos están en formato correcto para pandas
                                try:
                                    df = pd.DataFrame(data)
                                    df.to_excel(writer, sheet_name=table, index=False)
                                    tables_exported.append(table)
                                    excel_log.write(f"✓ Tabla {table} exportada exitosamente. {len(data)} registros.\n")
                                except Exception as df_error:
                                    excel_log.write(f"Error al convertir datos de {table} a DataFrame: {str(df_error)}\n")
                            else:
                                excel_log.write(f"La tabla {table} existe pero no contiene datos.\n")
                                # Add empty DataFrame to include the table structure
                                cursor.execute(f'DESCRIBE `{table}`')
                                columns = cursor.fetchall()
                                column_names = [col['Field'] for col in columns]
                                empty_df = pd.DataFrame(columns=column_names)
                                empty_df.to_excel(writer, sheet_name=table, index=False)
                                tables_exported.append(table)
                                excel_log.write(f"Se ha exportado la estructura de la tabla {table} sin datos.\n")
                        except Exception as query_error:
                            excel_log.write(f"Error al ejecutar consulta en tabla {table}: {str(query_error)}\n")
                    else:
                        excel_log.write(f"La tabla {table} no fue encontrada en la base de datos.\n")
                        
                        # Si la tabla es 'delivery', imprimir información adicional para diagnóstico
                        if table == 'delivery':
                            excel_log.write("ATENCIÓN: La tabla 'delivery' no existe o no es accesible\n")
                            # Verificar si podemos enumerar tablas para diagnóstico
                            try:
                                cursor.execute("SHOW TABLES")
                                tables_in_db = [row[f"Tables_in_{app.config['MYSQL_DB']}"] for row in cursor.fetchall()]
                                excel_log.write(f"Tablas disponibles en la base de datos: {', '.join(tables_in_db)}\n")
                            except Exception as list_error:
                                excel_log.write(f"No se pudieron listar las tablas: {str(list_error)}\n")
                except Exception as table_error:
                    excel_log.write(f"Error al exportar la tabla {table}: {str(table_error)}\n")

        # Guardar el archivo Excel
        try:
            # Verificar si hay tablas para exportar
            if not tables_exported:
                with open(excel_log_path, 'a', encoding='utf-8') as excel_log:
                    excel_log.write("ADVERTENCIA: No se ha exportado ninguna tabla. Verificando tablas en la base de datos...\n")
                    # Intentar obtener todas las tablas de la base de datos
                    cursor.execute("SHOW TABLES")
                    all_tables = [row[f"Tables_in_{app.config['MYSQL_DB']}"] for row in cursor.fetchall()]
                    excel_log.write(f"Tablas encontradas en la base de datos: {', '.join(all_tables)}\n")
                    excel_log.write("Intentando exportar todas las tablas disponibles...\n")
                    
                    # Intentar exportar todas las tablas encontradas
                    for db_table in all_tables:
                        try:
                            cursor.execute(f"SELECT * FROM `{db_table}`")
                            data = cursor.fetchall()
                            if data:
                                df = pd.DataFrame(data)
                                df.to_excel(writer, sheet_name=db_table[:31], index=False)  # Excel limit sheet name to 31 chars
                                tables_exported.append(db_table)
                                excel_log.write(f"Tabla {db_table} exportada exitosamente como respaldo. {len(data)} registros.\n")
                        except Exception as table_error:
                            excel_log.write(f"Error al exportar tabla {db_table} como respaldo: {str(table_error)}\n")
            
            writer.close()
            with open(excel_log_path, 'a', encoding='utf-8') as excel_log:
                excel_log.write(f"Excel guardado correctamente en {file_path}\n")
                excel_log.write(f"Tablas exportadas: {', '.join(tables_exported)}\n")
                
                # Verificar si se exportó la tabla 'delivery'
                if 'delivery' not in tables_exported:
                    excel_log.write("ADVERTENCIA: La tabla 'delivery' no se ha exportado. Verifique los permisos y la existencia de la tabla.\n")
        except Exception as save_error:
            with open(excel_log_path, 'a', encoding='utf-8') as excel_log:
                excel_log.write(f"Error al guardar el archivo Excel: {str(save_error)}\n")
            raise

        # Registrar la acción en el historial
        cursor.execute(
            'INSERT INTO user_history (cedula, user, action, time_login) VALUES (%s, %s, %s, %s)',
            (session['cedula'], session['username'], 'Generó backup Excel de la base de datos', datetime.now())
        )
        MySQL.connection.commit()
        cursor.close()

        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        if 'cursor' in locals():
            cursor.close()
        error_msg = str(e)
        
        # Registrar el error en un archivo de log para debugging
        try:
            backup_dir = os.path.join(app.root_path, "backups")
            os.makedirs(backup_dir, exist_ok=True)
            error_log = os.path.join(backup_dir, f"excel_error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
            
            with open(error_log, 'w', encoding='utf-8') as f:
                f.write(f"Error timestamp: {datetime.now()}\n")
                f.write(f"Error message: {error_msg}\n")
                f.write(f"Error type: {type(e).__name__}\n")
                
                # Si es un error de MySQL, intentar obtener más detalles
                if hasattr(e, 'args') and len(e.args) > 0:
                    f.write(f"Error details: {e.args}\n")
        except Exception as log_error:
            # Si no podemos registrar el error, al menos añadirlo al mensaje
            error_msg += f" (Error adicional al registrar log: {str(log_error)})"
        
        return render_template("gestionar_data.html", error=f"Error al crear backup Excel: {error_msg}")


@app.route("/backup_excel", methods=["GET", "POST"])
def backup_excel_route():
    """Ruta para generar backup Excel de la base de datos."""
    # Verificar autenticación primero
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # Verificar permisos
    if session.get('Super_Admin') != 1:
        return render_template("gestionar_data.html", error="No tiene permisos para realizar esta acción.")
    
    try:
        # Manejo según el método HTTP
        if request.method == "POST":
            # Verificar si es una confirmación de backup
            action = request.form.get('action')
            if action == 'confirm':
                # Generar el backup Excel
                return backup_to_excel()
            else:
                # Si no hay acción o es diferente, redirigir a la página principal
                return redirect(url_for('gestionar_data'))
        
        elif request.method == "GET":
            # Mostrar la página de confirmación en la interfaz
            return render_template("gestionar_data.html")
        
        else:
            # Método no permitido (diferente de GET o POST)
            return render_template("gestionar_data.html", error="Método no permitido para esta operación."), 405
    
    except Exception as e:
        # Capturar cualquier error no manejado
        return render_template("gestionar_data.html", error=f"Error al generar backup Excel: {str(e)}")


# Ruta para gestionar la DB (unifica las funciones de backup y carga)
@app.route("/gestionar_data", methods=["GET", "POST"])
def gestionar_data():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    if session.get('Super_Admin') != 1:
        return render_template("gestionar_data.html", error="No tiene permisos para realizar esta acción.")
    
    if request.method == "POST":
        # Procesar formularios para confirmación de backups
        action = request.form.get('action')
        tipo_backup = request.form.get('tipo_backup')
        
        try:
            if action == 'confirm' and tipo_backup == 'sql':
                return backup_to_sql()
            elif action == 'confirm' and tipo_backup == 'excel':
                return backup_to_excel()
            elif 'file' in request.files:
                # Si se está subiendo un archivo, manejar la carga
                return cargar_data()
        except Exception as e:
            return render_template("gestionar_data.html", error=f"Error al procesar la solicitud: {str(e)}")
    
    # Para solicitudes GET, simplemente mostrar la página de gestión
    return render_template("gestionar_data.html")
def check_session():
    if 'loggedin' in session:
        return jsonify({"active": True})
    else:
        return jsonify({"active": False})

if __name__ == '__main__':
    app.run(host='0.0.0.0')
    